# import tkinter as tk
# from tkinter import filedialog
# import openpyxl
# from googlesearch import search
# import re
# import spacy

# class BusinessInfoGUI:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("Business Information Gatherer")

#         self.load_excel_button = tk.Button(self.root, text="Load Excel File", command=self.load_excel)
#         self.load_excel_button.pack()

#         self.input_label = tk.Label(self.root, text="Enter Business Name:")
#         self.input_label.pack()

#         self.input_entry = tk.Entry(self.root)
#         self.input_entry.pack()

#         self.start_button = tk.Button(self.root, text="Start Gathering", command=self.start_gathering)
#         self.start_button.pack()

#         self.nlp = spacy.load("en_core_web_sm")

#     def load_excel(self):
#         self.excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
#         self.column_headers = self.get_column_headers()

#     def get_column_headers(self):
#         # Retrieve column headers from the second row of the Excel file
#         wb = openpyxl.load_workbook(self.excel_file_path)
#         ws = wb.active
#         headers = []
#         for cell in ws[2]:
#             headers.append(cell.value)
#         return headers

#     def start_gathering(self):
#         # Get business name from input
#         input_text = self.input_entry.get()
#         business_name = self.extract_business_name(input_text)

#         # Determine the required information from column headers
#         required_info = self.get_required_info()

#         # Scrape data from the internet based on required information
#         business_info = self.scrape_business_info(business_name, required_info)

#         # Update Excel sheet with scraped data
#         self.update_excel_sheet(business_info)

#     def extract_business_name(self, text):
#         # Use spaCy for more sophisticated NLP-based entity recognition
#         doc = self.nlp(text)
#         for ent in doc.ents:
#             if ent.label_ == "ORG":
#                 return ent.text
#         return "Business"  # Default name if no organization entity found

#     def get_required_info(self):
#         # Use NLP to extract keywords related to required information from column headers
#         required_info = []
#         for header in self.column_headers:
#             keywords = re.findall(r'\b(address|phone|email|website)\b', header.lower())
#             required_info.extend(keywords)
#         return list(set(required_info))  # Remove duplicates and convert to list

#     def scrape_business_info(self, business_name, required_info):
#         # Construct search query
#         query = f'{business_name} {" ".join(required_info)}'

#         # Scrape data from search results
#         business_info = {}
#         for result in search(query, num=1, stop=len(required_info), pause=2):
#             for info_type in required_info:
#                 if info_type in result:
#                     business_info[info_type] = result
#                     break

#         return business_info

#     def update_excel_sheet(self, business_info):
#         # Load existing data from Excel file
#         wb = openpyxl.load_workbook(self.excel_file_path)
#         ws = wb.active

#         # Find the next available row
#         last_row = ws.max_row
#         row = last_row + 1

#         # Write scraped data to Excel sheet
#         for col, info_type in enumerate(business_info.keys(), start=1):
#             ws.cell(row=row, column=col, value=business_info[info_type])

#         # Save updated Excel file
#         wb.save(self.excel_file_path)

# if __name__ == "__main__":
#     root = tk.Tk()
#     app = BusinessInfoGUI(root)
#     root.mainloop()
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from googlesearch import search
import re
import spacy
import requests
from bs4 import BeautifulSoup
import os

class BusinessInfoGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Business Information Gatherer")

        self.load_excel_button = tk.Button(self.root, text="Load Excel File", command=self.load_excel)
        self.load_excel_button.pack()

        self.input_label = tk.Label(self.root, text="Enter Business Name:")
        self.input_label.pack()

        self.input_entry = tk.Entry(self.root)
        self.input_entry.pack()

        self.start_button = tk.Button(self.root, text="Start Gathering", command=self.start_gathering)
        self.start_button.pack()

        self.nlp = spacy.load("en_core_web_sm")
        self.excel_file_path = None
        self.column_headers = []

    def load_excel(self):
        self.excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.excel_file_path:
            self.column_headers = self.get_column_headers()
            if not self.column_headers:
                messagebox.showerror("Error", "No headers found in the Excel sheet!")
        else:
            messagebox.showerror("Error", "No file selected!")

    def get_column_headers(self):
        wb = openpyxl.load_workbook(self.excel_file_path)
        ws = wb.active
        headers = []
        try:
            for cell in ws[2]:
                if cell.value:
                    headers.append(cell.value)
        except IndexError:
            messagebox.showerror("Error", "The sheet does not have enough rows.")
        return headers

    def start_gathering(self):
        if not self.excel_file_path:
            messagebox.showerror("Error", "Please load an Excel file first.")
            return

        input_text = self.input_entry.get().strip()
        if not input_text:
            messagebox.showerror("Error", "Please enter a valid business name.")
            return

        business_name = self.extract_business_name(input_text)
        required_info = self.get_required_info()

        if not required_info:
            messagebox.showerror("Error", "No valid headers found in the Excel sheet.")
            return

        business_info = self.scrape_business_info(business_name, required_info)

        if not business_info:
            messagebox.showerror("Error", "Failed to scrape business information.")
            return

        self.update_excel_sheet(business_info)
        self.open_excel_file()

    def extract_business_name(self, text):
        doc = self.nlp(text)
        for ent in doc.ents:
            if ent.label_ == "ORG":
                return ent.text
        return text

    def get_required_info(self):
        required_info = []
        for header in self.column_headers:
            keywords = re.findall(
                r'\b(address|phone|email|website|logo|image|description|contact|social media|company name|city|country|zip code)\b',
                header.lower())
            required_info.extend(keywords)
        return list(set(required_info))

    def scrape_business_info(self, business_name, required_info):
        query = f'{business_name} {" ".join(required_info)}'
        business_info = {}
        try:
            search_results = search(query, num_results=10)
            for result in search_results:
                response = requests.get(result)
                if response.status_code == 200:
                    soup = BeautifulSoup(response.text, 'html.parser')

                    # Remove script and style tags
                    for script in soup(["script", "style"]):
                        script.decompose()

                    for info_type in required_info:
                        if info_type == "address" and not business_info.get("address"):
                            address = soup.find(string=re.compile(r'\bAddress\b', re.I))
                            if address:
                                business_info["address"] = address.find_next().get_text(strip=True)
                        elif info_type == "phone" and not business_info.get("phone"):
                            phone = soup.find(string=re.compile(r'\bPhone\b', re.I))
                            if phone:
                                business_info["phone"] = phone.find_next().get_text(strip=True)
                        elif info_type == "email" and not business_info.get("email"):
                            email = soup.find(string=re.compile(r'\bEmail\b', re.I))
                            if email:
                                business_info["email"] = email.find_next().get_text(strip=True)
                        elif info_type == "website" and not business_info.get("website"):
                            website = soup.find("a", href=True, string=re.compile(r'\bWebsite\b', re.I))
                            if website:
                                business_info["website"] = website['href']
                        elif (info_type == "logo" or info_type == "image") and not business_info.get("image"):
                            image = soup.find("img", src=True)
                            if image:
                                image_url = image['src']
                                if image_url.startswith('http'):
                                    business_info["image"] = image_url
                                else:
                                    business_info["image"] = response.url + image_url
                        elif info_type == "description" and not business_info.get("description"):
                            description = soup.find("meta", {"name": "description"})
                            if description:
                                business_info["description"] = description['content']
                        elif info_type == "contact" and not business_info.get("contact"):
                            contact = soup.find(string=re.compile(r'\bContact\b', re.I))
                            if contact:
                                business_info["contact"] = contact.find_next().get_text(strip=True)
                        elif info_type == "social media" and not business_info.get("social media"):
                            social_media = soup.find("a", href=re.compile(r'(facebook|twitter|instagram|linkedin)', re.I))
                            if social_media:
                                business_info["social media"] = social_media['href']
                        elif info_type == "company name" and not business_info.get("company name"):
                            company_name = soup.find(string=re.compile(r'\bCompany Name\b', re.I))
                            if company_name:
                                business_info["company name"] = company_name.find_next().get_text(strip=True)
                        elif info_type == "city" and not business_info.get("city"):
                            city = soup.find(string=re.compile(r'\bCity\b', re.I))
                            if city:
                                business_info["city"] = city.find_next().get_text(strip=True)
                        elif info_type == "country" and not business_info.get("country"):
                            country = soup.find(string=re.compile(r'\bCountry\b', re.I))
                            if country:
                                business_info["country"] = country.find_next().get_text(strip=True)
                        elif info_type == "zip code" and not business_info.get("zip code"):
                            zip_code = soup.find(string=re.compile(r'\bZip Code\b', re.I))
                            if zip_code:
                                business_info["zip code"] = zip_code.find_next().get_text(strip=True)
                    
                    # Break loop if all required info is gathered
                    if len(business_info) == len(required_info):
                        break
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during scraping: {e}")
            return None
        return business_info

    def update_excel_sheet(self, business_info):
        wb = openpyxl.load_workbook(self.excel_file_path)
        ws = wb.active

        # Find the header row index
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                if cell.value in self.column_headers:
                    header_row = cell.row
                    break
            if header_row:
                break

        if header_row is None:
            messagebox.showerror("Error", "No matching headers found in the Excel sheet!")
            return

        # Write data into corresponding columns
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if header in self.column_headers:
                info_type = header.lower()
                value = business_info.get(info_type, "")
                # Find the next empty cell in the column
                for row in range(header_row + 1, ws.max_row + 2):
                    if ws.cell(row=row, column=col).value is None:
                        ws.cell(row=row, column=col, value=value)
                        break

        wb.save(self.excel_file_path)
        messagebox.showinfo("Success", "Business information updated successfully!")

    def open_excel_file(self):
        try:
            os.startfile(self.excel_file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel file: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = BusinessInfoGUI(root)
    root.mainloop()
